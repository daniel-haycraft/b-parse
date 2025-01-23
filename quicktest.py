import openpyxl
from openpyxl.utils import column_index_from_string
from datetime import datetime
from collections import defaultdict

def number_entries_by_date(file_path, date_column="R", output_column="D", start_row=2, end_row=2834):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Convert column letters to indices
    date_col_idx = column_index_from_string(date_column)
    output_col_idx = column_index_from_string(output_column)
    
    # Dictionary to store dates by value in column A
    date_rows = defaultdict(list)
    
    # Collect rows with dates for each unique value in column A
    for row in range(start_row, end_row + 1):  # Loop through column A from specified start to end row
        cell_value = sheet.cell(row=row, column=1).value  # Only look in column A
        date_cell = sheet.cell(row=row, column=date_col_idx).value
        
        # Only consider rows with a non-empty value in column A and a valid date in column R
        if cell_value is not None and isinstance(date_cell, datetime):
            date_rows[cell_value].append((row, date_cell))
            print(f"Row {row} with value '{cell_value}' added with date {date_cell}")

    # Sort and number rows for each unique value
    for value, rows in date_rows.items():
        # Sort rows by date in ascending order
        sorted_rows = sorted(rows, key=lambda x: x[1])
        
        # Write sequential numbers in the output column based on sorted date order
        for index, (row, date) in enumerate(sorted_rows, start=1):
            sheet.cell(row=row, column=output_col_idx).value = index
            print(f"Row {row} for value '{value}' numbered {index}")

    # Save the workbook
    wb.save(file_path)
    print(f"Entries in column {output_column} numbered in chronological order by date for each value in column A.")

# Example usage
file_path = 'Copy of draws import.xlsx'
number_entries_by_date(file_path)
